#!/usr/bin/env python3
"""Fetch the cGSA supplementary workbooks and flatten them to committed TSV/GMT.

cGSA (Wang et al., *Bioinformatics* 2026, btag214; https://github.com/ncbi-nlp/cGSA)
ships its benchmark and results as Excel workbooks under ``Supplementary/``. This
script pulls those workbooks from ``raw.githubusercontent.com`` and writes the
flat, diff-friendly artifacts the eval adapters consume. Re-run to refresh.

Outputs (all under this directory):
  benchmark_sets.tsv        one row per benchmark DEG set (contexts + counts)
  benchmark_ground_truth.tsv  set_id, ground-truth pathway name (free text), rank
  benchmark_members.gmt     set_id -> DEG gene symbols (the enrichment input)
  case_study_sets.tsv       the 9 case-study DEG sets (metadata)
  case_study_members.gmt    case-study set_id -> gene symbols
  cgsa_ground_truth.tsv     result_id, pathway (PubMed ground truth), category
  cgsa_predictions.tsv      result_id, pathway (cGSA output), category, relevance

Nothing here is authored by this repo; it is a faithful re-serialization of the
cGSA authors' released data. See PROVENANCE.md.
"""

from __future__ import annotations

import argparse
import csv
import sys
import urllib.parse
import urllib.request
from pathlib import Path

HERE = Path(__file__).resolve().parent
RAW = "https://raw.githubusercontent.com/ncbi-nlp/cGSA/main/Supplementary"
WORKBOOKS = {
    "Curated DEGs.xlsx",
    "Results of cGSA.xlsx",
}


def fetch(name: str, dest_dir: Path) -> Path:
    dest = dest_dir / name
    if dest.exists():
        return dest
    url = f"{RAW}/{urllib.parse.quote(name)}"
    print(f"fetch {url}", file=sys.stderr)
    with urllib.request.urlopen(url) as resp:  # noqa: S310 (trusted host)
        dest.write_bytes(resp.read())
    return dest


def _split(value, sep: str) -> list[str]:
    if value is None:
        return []
    return [p.strip() for p in str(value).split(sep) if p.strip()]


def _write_tsv(path: Path, header: list[str], rows: list[list]) -> None:
    with path.open("w", newline="") as handle:
        writer = csv.writer(handle, delimiter="\t")
        writer.writerow(header)
        writer.writerows(rows)
    print(f"wrote {path.relative_to(HERE.parent.parent)}  ({len(rows)} rows)", file=sys.stderr)


def _write_gmt(path: Path, records: list[tuple[str, str, list[str]]]) -> None:
    with path.open("w") as handle:
        for set_id, desc, genes in records:
            handle.write("\t".join([set_id, desc, *genes]) + "\n")
    print(f"wrote {path.relative_to(HERE.parent.parent)}  ({len(records)} sets)", file=sys.stderr)


def build_benchmark(wb, out_dir: Path) -> None:
    ws = wb["Evaluated DEGs"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))  # row 1 is the title banner
    header = [h.strip() if isinstance(h, str) else h for h in rows[0]]
    idx = {h: i for i, h in enumerate(header)}
    data = rows[1:]

    set_rows: list[list] = []
    gt_rows: list[list] = []
    members: list[tuple[str, str, list[str]]] = []
    for r in data:
        set_id = str(r[idx["PMC"]]).strip()
        degs = _split(r[idx["DEGs"]], ",")
        gts = _split(r[idx["Pathway"]], ";")
        species = str(r[idx["Species (9606 for HOMO; 10090 for MUS)"]]).strip()
        set_rows.append([
            set_id,
            str(r[idx["Type"]] or "").strip(),
            f"NCBITaxon:{species}" if species else "",
            str(r[idx["Database"]] or "").strip(),
            str(r[idx["Research Objectives"]] or "").strip(),
            str(r[idx["Keywords (KW)"]] or "").strip(),
            str(r[idx["Curated Context (Refine)"]] or "").strip(),
            str(r[idx["Experimental Conditions"]] or "").strip(),
            str(r[idx["Source"]] or "").strip(),
            str(len(degs)),
            str(len(gts)),
        ])
        for rank, name in enumerate(gts, start=1):
            gt_rows.append([set_id, name, str(rank)])
        members.append((set_id, "cgsa_benchmark_degs", degs))

    _write_tsv(
        out_dir / "benchmark_sets.tsv",
        ["set_id", "type", "taxon", "database", "research_objective",
         "keywords", "curated_context", "experimental_conditions", "source",
         "n_degs", "n_ground_truth"],
        set_rows,
    )
    _write_tsv(out_dir / "benchmark_ground_truth.tsv",
               ["set_id", "ground_truth_pathway", "rank"], gt_rows)
    _write_gmt(out_dir / "benchmark_members.gmt", members)


def build_case_studies(wb, out_dir: Path) -> None:
    ws = wb["Case-Study DEGs"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    header = [h.strip() if isinstance(h, str) else h for h in rows[0]]
    idx = {h: i for i, h in enumerate(header)}
    data = rows[1:]

    set_rows: list[list] = []
    members: list[tuple[str, str, list[str]]] = []
    for r in data:
        set_id = str(r[idx["PMC"]]).strip()
        if not set_id or set_id == "None":
            continue
        genes = _split(r[idx["Gene List"]], " ")
        species = str(r[idx["Species (9606 for HOMO; 10090 for MUS)"]]).strip()
        set_rows.append([
            set_id,
            str(r[idx["Type"]] or "").strip(),
            f"NCBITaxon:{species}" if species else "",
            str(r[idx["Research Objectives"]] or "").strip(),
            str(r[idx["Keys"]] or "").strip(),
            str(r[idx["Experimental Conditions"]] or "").strip(),
            str(len(genes)),
        ])
        members.append((set_id, "cgsa_case_study_degs", genes))

    _write_tsv(
        out_dir / "case_study_sets.tsv",
        ["set_id", "type", "taxon", "research_objective", "keys",
         "experimental_conditions", "n_genes"],
        set_rows,
    )
    _write_gmt(out_dir / "case_study_members.gmt", members)


def build_results(wb, out_dir: Path) -> None:
    """Flatten the expert 'Functional Category Annotation' sheet.

    Within one result block, column C/D is the list of PubMed ground-truth
    pathways + expert category number, and column E/F is the parallel list of
    cGSA output pathways + category number. The two lists are NOT row-aligned
    pairs; the integer category is the linking key (matching category = a hit).
    A category of 0 means the expert judged the pathway unrelated to any
    ground-truth functional category.
    """
    ws = wb["Functional Category Annotation"]
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    data = rows[1:]  # row 0 is the header

    gt_rows: list[list] = []
    pred_rows: list[list] = []
    cur_id = None
    for r in data:
        _id, _ctx, gtp, gtc, cp, cc, rel = (list(r) + [None] * 7)[:7]
        if _id not in (None, ""):
            cur_id = str(_id).strip()
        if gtp not in (None, ""):
            gt_rows.append([cur_id, str(gtp).strip(), _catstr(gtc)])
        if cp not in (None, ""):
            pred_rows.append([cur_id, str(cp).strip(), _catstr(cc),
                              "" if rel in (None, "") else str(rel).strip()])

    _write_tsv(out_dir / "cgsa_ground_truth.tsv",
               ["result_id", "ground_truth_pathway", "category"], gt_rows)
    _write_tsv(out_dir / "cgsa_predictions.tsv",
               ["result_id", "cgsa_pathway", "category", "relevance_score"], pred_rows)


def _catstr(value) -> str:
    if value in (None, ""):
        return ""
    try:
        return str(int(float(value)))
    except (TypeError, ValueError):
        return str(value).strip()


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__,
                                     formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--src-dir", type=Path, default=HERE / "_workbooks",
                        help="Where the .xlsx are cached / downloaded to.")
    parser.add_argument("--out-dir", type=Path, default=HERE)
    args = parser.parse_args()

    try:
        import openpyxl  # noqa: F401
    except ImportError:
        print("openpyxl is required: pip install openpyxl", file=sys.stderr)
        return 1
    import openpyxl

    args.src_dir.mkdir(parents=True, exist_ok=True)
    args.out_dir.mkdir(parents=True, exist_ok=True)

    curated = openpyxl.load_workbook(fetch("Curated DEGs.xlsx", args.src_dir), data_only=True)
    results = openpyxl.load_workbook(fetch("Results of cGSA.xlsx", args.src_dir), data_only=True)

    build_benchmark(curated, args.out_dir)
    build_case_studies(curated, args.out_dir)
    build_results(results, args.out_dir)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
